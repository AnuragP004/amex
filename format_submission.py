import pandas as pd
import openpyxl
import shutil

print("Copying template to final_submission.xlsx...")
shutil.copy('6a3cb64c7cae4_campus_challenge_r1_submission_template.xlsx', 'final_submission.xlsx')

print("Reading scores...")
scores = pd.read_csv('submission.csv')
if 'id' in scores.columns:
    scores = scores.rename(columns={'id': 'ID'})
if 'profitability_score' in scores.columns:
    scores = scores.rename(columns={'profitability_score': 'Prediction'})

# Create a dictionary for fast lookup: { ID: Prediction }
score_dict = dict(zip(scores['ID'], scores['Prediction']))

print("Loading workbook with openpyxl to preserve all sheets...")
wb = openpyxl.load_workbook('final_submission.xlsx')
ws = wb['Predictions']

# Find which column is ID and which is Prediction
id_col_idx = None
pred_col_idx = None

for cell in ws[1]:
    if cell.value == 'ID':
        id_col_idx = cell.column
    elif cell.value == 'Prediction':
        pred_col_idx = cell.column

print(f"Found ID column at {id_col_idx}, Prediction column at {pred_col_idx}")

if id_col_idx and pred_col_idx:
    print("Populating predictions...")
    for row in range(2, ws.max_row + 1):
        id_val = ws.cell(row=row, column=id_col_idx).value
        if id_val in score_dict:
            ws.cell(row=row, column=pred_col_idx, value=score_dict[id_val])
else:
    print("Could not find ID or Prediction columns")

wb.save('final_submission.xlsx')
print("Successfully saved final_submission.xlsx with both 'Predictions' and 'Profitability Framework' sheets intact!")
